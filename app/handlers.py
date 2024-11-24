from aiogram import Router, F
from aiogram.filters import CommandStart
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
import asyncio
from io import StringIO, BytesIO                    # Для добавления новых слов 
from openpyxl import load_workbook                  # в категорию при помощи файла

from random import shuffle                          #Перемешивание слов, чтобы меньше работала зрительная память
from datetime import date, timedelta

import app.keyboards as kb
import app.database.requests as rq
import app.database.user_models as user_md



router = Router()


#region States

# Состояния для создания нового словаря
class RegistationNewDict(StatesGroup):
    name = State()         # Ожидание имени словаря
    matching = State()     # Ожидание соответствия


# Состояния для создания новой категории внутри словаря
class RegistationNewCat(StatesGroup):
    user_dict_id = ''      # Хранение ID словаря, к которому добавляется категория
    name = State()         # Ожидание имени категории


# Состояния для переключения страниц словарей
class UserDictsPages(StatesGroup):
    check = State()        # Флаг для проверки является ли нынешнее состояние листанием страниц словарей
    current_page = 0       # Индекс текущей страницы для отображения
    pages = []             # Список страниц словарей, содержащих по 10 словарей


# Состояния для переключения страниц категорий
class CategorisePages(StatesGroup):
    check = State()        # Флаг для проверки является ли нынешнее состояние листанием страниц категорий
    current_page = 0       # Индекс текущей страницы для отображения
    pages = []             # Список страниц категорий, содержащих по 10 категорий


class RepeatingWordsPages(StatesGroup):
    check = State()                    # Флаг для проверки является ли нынешнее состояние листанием страниц слов
    current_page = 0                   # Индекс текущей страницы для отображения слов
    pages = []                         # Список страниц слов, содержащих по 15 слов
    less_name = False                  # Опция: скрыть изучаемые слова
    less_matching = False              # Опция: скрыть соответствие слова


# Состояния для отображения и работы со словами
class WordsPages(StatesGroup):
    check = State()                    # Флаг для проверки является ли нынешнее состояние листанием страниц слов
    current_page = 0                   # Индекс текущей страницы для отображения слов
    pages = []                         # Список страниц слов, содержащих по 15 слов
    less_name = False                  # Опция: скрыть изучаемые слова
    less_matching = False              # Опция: скрыть соответствие слова
    less_common_words = False          # Опция: скрыть слова, не вызывающие затруднения у пользователя


# Состояния для создания нового элемента внутри категории
class RegistationNewItem(StatesGroup):
    category_id = ''       # ID категории, к которой добавляется элемент
    name = State()         # Ожидание имени элемента от пользователя
    matching = State()     # Ожидание соответствия элемента


# Состояния для подтверждения удаления элемента из категории
class ConfirmDelWord(StatesGroup):
    category_id = ''       # ID категории, из которой удаляется элемент
    name = State()         # Имя элемента, который нужно удалить
    matching = State()     # Соответствие элемента, который нужно удалить
    confirm = State()      # Подтверждение удаления


# Состояния для режима обучения
class LearnWords(StatesGroup):
    categories_id = set()             # Список ID категорий для обучения
    user_dict_id = 0                  # ID выбранного пользователем словаря
    words = []                        # Список слов для обучения
    order_difficult_words = []        # Порядок сложных слов для приоритетного обучения
    last_word = {}                    # Последнее изучаемое слово
    name = State()                    # Ожидание изучаемого слова от пользователя
    id_message_for_delete = 0         # ID сообщения, которое необходимо удалить


# Состояния для загрузки нового файла со словами
class RegistationNewWords(StatesGroup):
    category_id = ''         # ID категории, в которую загружается файл
    file = State()           # Ожидание загрузки файла от пользователя

#endregion



####################################################################################

#                              РАБОТА С СЛОВАРЯМИ

####################################################################################
#region DICTIONARIES    


#ВЫВОД ПОЛЬЗОВАТЕЛЬСКИХ СЛОВАРЕЙ
#region OUTPUT DICTIONARIES


# Создание страниц словарей для отображения
async def base_for_display_user_dicts(user_id: int, state: FSMContext):

    pages = [[]]
    # Разбивка категорий на страницы по 10 элементов
    for item in await rq.get_user_dicts(user_id):
        if len(pages[-1]) == 10:
            pages.append([])
        pages[-1].append((item.id, item.name))
    
    # Установка состояния и сохранение текущей страницы
    await state.set_state(UserDictsPages.check)
    await state.update_data(pages=pages)
    await state.update_data(current_page=0)


# Обработчик команды /start
@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext, new_user: bool=True):

    # Получение ID пользователя, инициировавшего запрос
    user_id = message.from_user.id
    if isinstance(message, CallbackQuery):
        message = message.message

    # Проверка, является ли пользователь новым
    if new_user and not (await rq.is_new_user(user_id)).first():
        user_name = message.from_user.username

        # Создание базы данных для нового пользователя
        await user_md.create_user_database(user_id)
        await rq.add_new_user(user_name, user_id)

    # Отправка пользователю списка словарей для выбора

    if await state.get_state() != CategorisePages.check:
        await state.clear()
        await base_for_display_user_dicts(user_id, state)
    
    #Получение данных для выводы страниц словарей
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']

    await message.answer(
        f'Выбери словарь:',
        reply_markup=await kb.inline_dictionaries(pages[current_page], current_page, len(pages)),
        parse_mode='html'
    )


# Функция для вывода списка словарей пользователю
# Аналогична предыдущей, но вызывается с callback
# Отображение словарей
@router.callback_query(F.data == 'main')
async def display_user_dicts(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id

    # Инициализация состояния при первом вызове
    if await state.get_state() != UserDictsPages.check:
        await state.clear()
        await base_for_display_user_dicts(user_id, state)
    
    #Получение данных для выводы страниц категорий
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']

    await callback.message.edit_text(
        f'Выбери словарь:',
        reply_markup=await kb.inline_dictionaries(pages[current_page], current_page, len(pages)),
        parse_mode='html'
    )


# Переход к следующей странице словарей
@router.callback_query(F.data.startswith('next dict page'))
async def next_user_dicts_page(callback: CallbackQuery, state: FSMContext):

    #Увеличили индекс страницы на единицу. Функция не будет доступна, если индекс максимален
    data = await state.get_data()
    await state.update_data(current_page=data['current_page'] + 1)

    await display_user_dicts(callback, state)


# Переход к предыдущей странице словарей
@router.callback_query(F.data.startswith('previous dict page'))
async def previous_user_dicts_page(callback: CallbackQuery, state: FSMContext):

    #Уменьшили индекс страницы на единицу. Функция не будет доступна, если индекс равен 0
    data = await state.get_data()
    await state.update_data(current_page=data['current_page'] - 1)

    await display_user_dicts(callback, state)

#endregion



#СОЗДАНИЕ НОВОГО СЛОВАРЯ
#region
#-----------------------------------------------------------------------------------
#                          /start/Добавить новый словарь
#-----------------------------------------------------------------------------------


# Обработчик для добавления нового словаря через нажатие кнопки
@router.callback_query(F.data == 'add new dict')
async def add_new_dict(callback: CallbackQuery, state: FSMContext):
    
    # Удаление предыдущего сообщения, чтобы очистить интерфейс
    await callback.message.delete()

    # Проверка текущего состояния FSMContext (если уже существует, то очищаем)
    state_name = await state.get_state()
    if state_name is not None:
        await state.clear()

    # Установка нового состояния для ввода названия словаря
    await state.set_state(RegistationNewDict.name)

    await callback.message.answer('Введите название нового словаря',
                                  reply_markup=await kb.reply_cancel_add_new_item())


# Получение названия словаря
@router.message(RegistationNewDict.name)
async def set_name_dict(message: Message, state: FSMContext):

    # Проверка, отменил ли пользователь создание нового словаря
    if message.text == 'ОТМЕНА':
        return await cmd_start(message, state, new_user=False)  # Возврат к выбору словаря

    # Установка нового состояния для ввода соответствия словаря
    await state.set_state(RegistationNewDict.matching)

    # Сохранение введенного названия словаря в контексте состояния FSM
    await state.update_data(name=message.text)

    await message.answer('Введите соответствие', reply_markup=await kb.reply_cancel_add_new_item())


# Получение соответствия словаря
@router.message(RegistationNewDict.matching)
async def set_matching_dict(message: Message, state: FSMContext):

    user_id = message.from_user.id

    # Проверка, отменил ли пользователь создание нового словаря
    if message.text == 'ОТМЕНА':
        return await cmd_start(message, state, new_user=False)  # Возврат к выбору словаря

    # Сохранение введенного соответствия словаря в контексте состояния FSM
    await state.update_data(matching=message.text)
    data = await state.get_data()       # Получение всех сохраненных данных
    await state.clear()                 # Очистка состояния после ввода данных

    name_user_dict = data['name']
    matching_user_dict = data['matching']

    # Запись нового словаря в базе данных пользователя
    await rq.add_new_user_dict(user_id, name_user_dict, matching_user_dict)

    # Подтверждение создания словаря пользователю
    await message.answer(f'Словарь <b>{name_user_dict}</b> создан!',
                         parse_mode="html")

    user_dict_id = (await rq.get_id_user_dict_by_name(user_id, name_user_dict)).first()

    await message.answer(f'Словарь <b>{name_user_dict}</b>',
                         reply_markup=await kb.inline_categories(user_dict_id, [], 0, 1, ),
                         parse_mode='html')



#-----------------------------------------------------------------------------------
#endregion



#РЕДАКТИРОВАНИЕ СЛОВАРЯ
#region
#-----------------------------------------------------------------------------------
#                     /start/{название словаря}/редактировать словарь
#-----------------------------------------------------------------------------------


# Обработчик для редактирования словаря
@router.callback_query(F.data.startswith('edit dict'))
async def edit_user_dict(callback: CallbackQuery):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "edit dict_{user_dict_id}"
    
    user_id = callback.from_user.id
    user_dict_id = callback.data.split('_')[1]

    name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()

    await callback.message.edit_text(f'Словарь <b>{name_user_dict}</b>',
                                     reply_markup=await kb.inline_edit_user_dict(user_dict_id),
                                     parse_mode='html')



#-----------------------------------------------------------------------------------
#endregion



#УДАЛЕНИЕ СЛОВАРЯ
#region
#-----------------------------------------------------------------------------------
#            /start/{название словаря}/редактировать словарь/удалить словарь
#-----------------------------------------------------------------------------------


# Обработчик для запроса на удаление словаря
@router.callback_query(F.data.startswith('del dict'))
async def del_user_dict(callback: CallbackQuery):
    
    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "del dict_{user_dict_id}"

    user_id = callback.from_user.id
    user_dict_id = callback.data.split('_')[1]

    name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()

    # Предложение пользователю подтвердить удаление словаря
    await callback.message.edit_text(
        f'Вы точно хотите удалить словарь {name_user_dict}?',
        reply_markup=await kb.inline_confirm_del_user_dict(user_dict_id)
    )


# Обработчик для подтверждения удаления словаря
@router.callback_query(F.data.startswith('confirm del dict'))
async def confirm_del_user_dict(callback: CallbackQuery, state: FSMContext):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "confirm del dict_{user_dict_id}"
    
    user_id = callback.from_user.id
    user_dict_id = callback.data.split('_')[1]

    name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()

    # Удаление словаря из базы данных
    await rq.delete_dict(user_id, user_dict_id)

    # Подтверждение пользователю об успешном удалении словаря
    await callback.message.edit_text(f"Словарь <b>{name_user_dict}</b> удалён", parse_mode='html')

    await display_user_dicts(callback, state)


#-----------------------------------------------------------------------------------
#endregion


#endregion





####################################################################################

#                              РАБОТА С КАТЕГОРИЕЙ

####################################################################################
#region CATEGORIES


#ВЫВОД КАТЕГОРИЙ СЛОВАРЯ
#region
#-----------------------------------------------------------------------------------
#                              /start/{название словаря}
#-----------------------------------------------------------------------------------


# Создание страниц категорий для отображения
async def base_for_display_categories(user_id: int, user_dict_id: int, state: FSMContext):

    pages = [[]]
    # Разбивка категорий на страницы по 10 элементов
    for item in await rq.get_categories(user_id, user_dict_id):
        if len(pages[-1]) == 10:
            pages.append([])
        pages[-1].append((item.id, item.name))
    
    # Установка состояния и сохранение текущей страницы
    await state.set_state(CategorisePages.check)
    await state.update_data(pages=pages)
    await state.update_data(current_page=0)


# Отображение категорий словаря
@router.callback_query(F.data.startswith('dict'))
async def display_categories(callback: CallbackQuery, state: FSMContext):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "dict_{user_dict_id}"

    user_id = callback.from_user.id
    user_dict_id = callback.data.split('_')[1]

    # Инициализация состояния при первом вызове
    if await state.get_state() != CategorisePages.check:
        await state.clear()
        await base_for_display_categories(user_id, user_dict_id, state)
    
    #Получение данных для выводы страниц категорий
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']

    name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()
    await callback.message.edit_text(
        f'Словарь <b>{name_user_dict}</b>',
        reply_markup=await kb.inline_categories(user_dict_id, pages[current_page], current_page, len(pages)),
        parse_mode='html'
    )


# Переход к следующей странице категорий
@router.callback_query(F.data.startswith('next cat page'))
async def next_categories_page(callback: CallbackQuery, state: FSMContext):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "next cat page_{user_dict_id}"

    #Увеличили индекс страницы на единицу. Функция не будет доступна, если индекс максимален
    data = await state.get_data()
    await state.update_data(current_page=data['current_page'] + 1)

    await display_categories(callback, state)


# Переход к предыдущей странице категорий
@router.callback_query(F.data.startswith('previous cat page'))
async def previous_categories_page(callback: CallbackQuery, state: FSMContext):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "previous cat page_{user_dict_id}"

    #Уменьшили индекс страницы на единицу. Функция не будет доступна, если индекс равен 0
    data = await state.get_data()
    await state.update_data(current_page=data['current_page'] - 1)

    await display_categories(callback, state)



#-----------------------------------------------------------------------------------
#endregion



#СОЗДАНИЕ НОВОЙ КАТЕГОРИИ
#region
#-----------------------------------------------------------------------------------
#    /start/{название словаря}/редактировать словарь/добавить новую категорию
#-----------------------------------------------------------------------------------


# Начало процесса добавления новой категории
@router.callback_query(F.data.startswith('add new cat'))
async def add_new_category(callback: CallbackQuery, state: FSMContext):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "add new cat_{user_dict_id}"

    user_id = callback.from_user.id
    user_dict_id = callback.data.split('_')[1]

    await callback.message.delete()

    # Очистка состояния, если оно установлено
    state_name = await state.get_state()
    if state_name is not None:
        await state.clear()

    await state.set_state(RegistationNewCat.name)

    await state.update_data(user_dict_id=user_dict_id)

    name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()

    await callback.message.answer(
        f'Введите название новой категории словаря <b>{name_user_dict}</b>',
        reply_markup=await kb.reply_cancel_add_new_item(),
        parse_mode='html'
    )


# Получение и обработка названия категории
@router.message(RegistationNewCat.name)
async def set_name_category(message: Message, state: FSMContext):

    user_id = message.from_user.id

    # Проверка на отмену процесса
    if message.text == 'ОТМЕНА':
        data = await state.get_data()
        await state.clear()
        user_dict_id = data['user_dict_id']
        name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()

        # Возврат к меню редактирования словаря
        await message.answer(
            f'Словарь <b>{name_user_dict}</b>',
            reply_markup=await kb.inline_edit_user_dict(user_dict_id),
            parse_mode='html'
        )
        return

    await state.update_data(name=message.text)
    data = await state.get_data()
    await state.clear()

    # Сохранение категории в базе данных и создание пустой страницы
    await rq.add_new_category(user_id, data['user_dict_id'], message.text)

    category_id = (await rq.get_id_category_by_name(user_id, message.text)).first()

    # Инициализация состояния для страницы слов
    pages = [[]]
    await state.set_state(WordsPages.check)
    await state.update_data(pages=pages)
    await state.update_data(current_page=0)
    await state.update_data(less_name=False)
    await state.update_data(less_matching=False)
    await state.update_data(less_common_words=False)

    await message.answer(
        f"Категория <b>{message.text}</b>\n\nЗдесь пока пусто",
        reply_markup=await kb.inline_words(
            user_id=user_id,
            category_id=category_id,
            user_dict_id=data['user_dict_id'],
            current_page=0,
            cnt_pages=len(pages),
            is_not_empty=False,
        ),
        parse_mode='html'
    )


#-----------------------------------------------------------------------------------
#endregion



#РЕДАКТИРОВАНИЕ КАТЕГОРИИ
#region
#-----------------------------------------------------------------------------------
#     /start/{название словаря}/{название категории}/редактировать категорию
#-----------------------------------------------------------------------------------


# Обработчик для редактирования категории
@router.callback_query(F.data.startswith('edit cat'))
async def edit_category(callback: CallbackQuery):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "edit cat_{category_id}"

    user_id = callback.from_user.id

    category_id = callback.data.split('_')[1]
    name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()

    await callback.message.edit_text(f'Категория <b>{name_category}</b>',
                                  reply_markup= await kb.inline_edit_category(category_id),
                                  parse_mode='html'
                                  )

#-----------------------------------------------------------------------------------
#endregion



#УДАЛЕНИЕ КАТЕГОРИИ
#region
#-----------------------------------------------------------------------------------
#          /start/{название словаря}/редактировать словарь/удалить словарь
#-----------------------------------------------------------------------------------


# Начало процесса удаления категории с подтверждением
@router.callback_query(F.data.startswith('del cat'))
async def del_category(callback: CallbackQuery):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "del cat_{category_id}"

    user_id = callback.from_user.id

    category_id = callback.data.split('_')[1]
    name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()

    # Запрос на подтверждение удаления категории
    await callback.message.edit_text(
        f'Вы точно хотите удалить категорию {name_category}?',
        reply_markup=await kb.inline_confirm_del_category(category_id)
    )


# Обработка подтверждения удаления категории
@router.callback_query(F.data.startswith('confirm del cat'))
async def confirm_del_category(callback: CallbackQuery):

    #ПРИМЕР ВХОДНЫХ ДАННЫХ
    #callback.data = "confirm del cat_{category_id}"

    user_id = callback.from_user.id

    # Получение данных о категории и словаре пользователя
    category_id = callback.data.split('_')[1]
    name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()
    name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()

    # Удаление категории и обновление интерфейса
    await rq.delete_category(user_id, category_id)
    await callback.message.edit_text(
        f"Категория <b>{name_category}</b> удалена",
        parse_mode='html'
    )

    await callback.message.answer(
        f'Словарь <b>{name_user_dict}</b>',
        reply_markup=await kb.inline_categories(user_dict_id, [], 0, 1),
        parse_mode='html'
    )


#-----------------------------------------------------------------------------------
#endregion


#endregion




####################################################################################

#                              РАБОТА СО СЛОВАМИ

####################################################################################
#region WORDS


#ВЫВОД СЛОВ СЛОВАРЯ
#region


# Инициализация списка слов для вывода на экране
async def base_for_display_repeating_words(user_id, user_dict_id, state):

    # Получаем слова категории и перемешиваем их
    items = [
        (item.id, item.name, item.matching)
        for item in await rq.get_repeating_words_by_dict(user_id, user_dict_id)
    ]
    shuffle(items)

    # Разбиваем слова на страницы по 15 элементов
    pages = []
    len_items = len(items)
    for num in range((len_items + 14) // 15):
        pages.append(items[num * 15: (num + 1) * 15])

    # Устанавливаем состояние и сохраняем данные о страницах и фильтрах
    await state.set_state(RepeatingWordsPages.check)
    await state.update_data(pages=pages)
    await state.update_data(current_page=0)
    await state.update_data(less_name=False)
    await state.update_data(less_matching=False)
    await state.update_data(less_common_words=False)


# Формирование текста для вывода слов
async def text_repeating_words(user_id, user_dict_id, pages, page, less_name, less_matching):

    name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()

    # Функция для форматирования каждого элемента списка
    def decorate(item, less_name, less_matching):
        if less_name == less_matching:
            name, matching = item
            return f"{name}  -  <i>{matching}</i>"
        if less_name:
            matching = item[0]
            return f"<i>{matching}</i>"
        name = item[0]
        return name

    # Возвращаем текст с отформатированным списком слов категории
    new_line = '\n'
    return f"Словарь <b>{name_user_dict}</b>\n\n{(new_line.join(decorate(item[1:], less_name, less_matching) for item in pages[page]) if pages else 'Здесь пока пусто')}"


# Обработчик вывода слов 
@router.callback_query(F.data.startswith('repeating words'))
async def display_repeating_words(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id

    user_dict_id = callback.data.split('_')[1]

    # Проверяем текущее состояние и инициализируем, если нужно
    name_state = await state.get_state()
    if name_state != RepeatingWordsPages.check:
        await state.clear()
        await base_for_display_repeating_words(user_id, user_dict_id, state)

    # Загружаем данные о текущей странице и фильтрах
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']
    less_name = data['less_name']
    less_matching = data['less_matching']

    await callback.message.edit_text(
        await text_repeating_words(
            user_id=user_id,
            user_dict_id=user_dict_id,
            pages=pages,
            page=current_page,
            less_name=less_name,
            less_matching=less_matching
            ),
        reply_markup=await kb.inline_repeating_words(
            user_id=user_id,
            user_dict_id=user_dict_id,
            current_page=current_page,
            cnt_pages=len(pages),
            is_not_empty=bool(pages[0] if pages else False),
            less_name = less_name,
            less_matching = less_matching
            ),
        parse_mode='html'
        )


# Переход к следующей странице
@router.callback_query(F.data.startswith('next repeating page'))
async def next_repeating_page(callback: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    current_page = data['current_page']

    #Увеличили индекс страницы на единицу. Функция не будет доступна, если индекс максимален
    await state.update_data(current_page=current_page + 1)

    await display_repeating_words(callback, state)


# Переход к предыдущей странице
@router.callback_query(F.data.startswith('previous repeating page'))
async def previous_repeating_page(callback: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    current_page = data['current_page']

    #Уменьшили индекс страницы на единицу. Функция не будет доступна, если индекс равен 0
    await state.update_data(current_page=current_page - 1)

    await display_repeating_words(callback, state)


# Удалить изучаемеые слова из отображения, оставить только соответствия
@router.callback_query(F.data.startswith('discard repeating name'))
async def discard_repeating_name(callback: CallbackQuery, state: FSMContext):

    # Получаем текущие страницы и удаляем слова, оставляя только соответствия
    data = await state.get_data()
    pages = data['pages']
    pages = [[(item[0], item[2]) for item in page] for page in pages]

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=pages)
    await state.update_data(less_name=True)
    await display_repeating_words(callback, state)


# Удалить соответствие из отображения, оставить только изучаемые слова
@router.callback_query(F.data.startswith('discard repeating matching'))
async def discard_repeating_matching(callback: CallbackQuery, state: FSMContext):

    # Получаем текущие страницы и удаляем соответствия, оставляя только изучаемые слова
    data = await state.get_data()
    pages = data['pages']
    pages = [[(item[0], item[1]) for item in page] for page in pages]

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=pages)
    await state.update_data(less_matching=True)
    await display_repeating_words(callback, state)


# Восстановить отображение слов
@router.callback_query(F.data.startswith('return repeating name'))
async def return_repeating_name(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id

    # Получаем исходные слова и добавляем их на страницы
    data = await state.get_data()
    pages = data['pages']
    new_pages = []
    for page in pages:
        new_pages.append([
            (item[0], (await rq.get_data_word_by_id(user_id, item[0], 'name')).first(), item[1])
            for item in page
        ])

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=new_pages)
    await state.update_data(less_name=False)
    await display_repeating_words(callback, state)


# Восстановить отображение соответствия
@router.callback_query(F.data.startswith('return repeating matching'))
async def return_repeating_matching(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id

    # Получаем исходные соответствия и добавляем их на страницы
    data = await state.get_data()
    pages = data['pages']
    new_pages = []
    for page in pages:
        new_pages.append([
            (item[0], item[1], (await rq.get_data_word_by_id(user_id, item[0], 'matching')).first())
            for item in page
        ])

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=new_pages)
    await state.update_data(less_matching=False)
    await display_repeating_words(callback, state)


# Перемешать порядок слов в текущей странице слов
@router.callback_query(F.data.startswith('shuffle repeating'))
async def shuffle_repeating_words(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id
    user_dict_id = callback.data.split('_')[1]

    # Получаем данные состояния и текст текущего сообщения
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']
    less_name = data['less_name']
    less_matching = data['less_matching']

    # Перемешиваем строки слов

    lines = callback.message.text.split('\n')

    if len(lines) >= 4:
        name_category = ' '.join(lines[0].split()[1:])
        items = pages[current_page]
        last_item = items.pop(-1)
        shuffle(items)

        items.insert(0, last_item)
        new_line = '\n'

        def decorate(item, less_name, less_matching):
            if less_name == less_matching:
                name, matching = item
                return f"{name}  -  <i>{matching}</i>"
            if less_name:
                matching = item[0]
                return f"<i>{matching}</i>"
            name = item[0]
            return name

        new_content = f"Словарь <b>{name_category}</b>\n\n{(new_line.join(decorate(item[1:], less_name, less_matching) for item in items))}"
        await callback.message.edit_text(
            new_content,
            reply_markup=await kb.inline_repeating_words(
                user_id=user_id,
                user_dict_id=user_dict_id,
                current_page=current_page,
                cnt_pages=len(pages),
                is_not_empty=bool(pages[0] if pages else False),
                less_name=less_name,
                less_matching=less_matching,
            ),
            parse_mode='html'
        )



#endregion


#ВЫВОД СЛОВ КАТЕГОРИИ
#region
#-----------------------------------------------------------------------------------
#                  /start/{название словаря}/{название категории}
#-----------------------------------------------------------------------------------


# Инициализация списка слов для вывода на экране
async def base_for_display_words(user_id, category_id, state):

    # Получаем слова категории и перемешиваем их
    items = [
        (item.id, item.name, item.matching)
        for item in await rq.get_words_by_category(user_id, category_id)
    ]
    shuffle(items)

    # Разбиваем слова на страницы по 15 элементов
    pages = []
    len_items = len(items)
    for num in range((len_items + 14) // 15):
        pages.append(items[num * 15: (num + 1) * 15])

    # Устанавливаем состояние и сохраняем данные о страницах и фильтрах
    await state.set_state(WordsPages.check)
    await state.update_data(pages=pages)
    await state.update_data(current_page=0)
    await state.update_data(less_name=False)
    await state.update_data(less_matching=False)
    await state.update_data(less_common_words=False)


# Формирование текста для вывода слов
async def text_words(user_id, category_id, pages, page, less_name, less_matching):

    name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()

    # Функция для форматирования каждого элемента списка
    def decorate(item, less_name, less_matching):
        if less_name == less_matching:
            name, matching = item
            return f"{name}  -  <i>{matching}</i>"
        if less_name:
            matching = item[0]
            return f"<i>{matching}</i>"
        name = item[0]
        return name

    # Возвращаем текст с отформатированным списком слов категории
    new_line = '\n'
    return f"Категория <b>{name_category}</b>\n\n{(new_line.join(decorate(item[1:], less_name, less_matching) for item in pages[page]) if pages else 'Здесь пока пусто')}"


# Обработчик вывода слов категории
@router.callback_query(F.data.startswith('cat'))
async def display_words(callback: CallbackQuery, state: FSMContext):

    # ПРИМЕР ВХОДНЫХ ДАННЫХ
    # callback.data = "cat_{category_id}"

    user_id = callback.from_user.id

    category_id = callback.data.split('_')[1]
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()

    # Проверяем текущее состояние и инициализируем, если нужно
    name_state = await state.get_state()
    if name_state != WordsPages.check:
        await state.clear()
        await base_for_display_words(user_id, category_id, state)

    # Загружаем данные о текущей странице и фильтрах
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']
    less_name = data['less_name']
    less_matching = data['less_matching']
    less_common_words = data['less_common_words']

    await callback.message.edit_text(
        await text_words(
            user_id=user_id,
            category_id=category_id,
            pages=pages,
            page=current_page,
            less_name=less_name,
            less_matching=less_matching
            ),
        reply_markup=await kb.inline_words(
            user_id=user_id,
            category_id=category_id,
            user_dict_id=user_dict_id,
            current_page=current_page,
            cnt_pages=len(pages),
            is_not_empty=bool(pages[0] if pages else False),
            less_name = less_name,
            less_matching = less_matching,
            less_common_words = less_common_words
            ),
        parse_mode='html'
        )


# Переход к следующей странице
@router.callback_query(F.data.startswith('next page'))
async def next_page(callback: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    current_page = data['current_page']

    #Увеличили индекс страницы на единицу. Функция не будет доступна, если индекс максимален
    await state.update_data(current_page=current_page + 1)

    await display_words(callback, state)


# Переход к предыдущей странице
@router.callback_query(F.data.startswith('previous page'))
async def previous_page(callback: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    current_page = data['current_page']

    #Уменьшили индекс страницы на единицу. Функция не будет доступна, если индекс равен 0
    await state.update_data(current_page=current_page - 1)

    await display_words(callback, state)


# Удалить изучаемеые слова из отображения категории, оставить только соответствия
@router.callback_query(F.data.startswith('discard name'))
async def discard_name(callback: CallbackQuery, state: FSMContext):

    # Получаем текущие страницы и удаляем слова, оставляя только соответствия
    data = await state.get_data()
    pages = data['pages']
    pages = [[(item[0], item[2]) for item in page] for page in pages]

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=pages)
    await state.update_data(less_name=True)
    await display_words(callback, state)


# Удалить соответствие из отображения категории, оставить только изучаемые слова
@router.callback_query(F.data.startswith('discard matching'))
async def discard_matching(callback: CallbackQuery, state: FSMContext):

    # Получаем текущие страницы и удаляем соответствия, оставляя только изучаемые слова
    data = await state.get_data()
    pages = data['pages']
    pages = [[(item[0], item[1]) for item in page] for page in pages]

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=pages)
    await state.update_data(less_matching=True)
    await display_words(callback, state)


# Восстановить отображение изучаемых слов в категории
@router.callback_query(F.data.startswith('return name'))
async def return_name(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id
    category_id = callback.data.split('_')[1]

    # Получаем исходные слова для текущей категории и добавляем их на страницы
    data = await state.get_data()
    pages = data['pages']
    new_pages = []
    for page in pages:
        new_pages.append([
            (item[0], (await rq.get_data_word_by_id(user_id, item[0], 'name')).first(), item[1])
            for item in page
        ])

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=new_pages)
    await state.update_data(less_name=False)
    await display_words(callback, state)


# Восстановить отображение соответствия в категории
@router.callback_query(F.data.startswith('return matching'))
async def return_matching(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id
    category_id = callback.data.split('_')[1]

    # Получаем исходные соответствия для текущей категории и добавляем их на страницы
    data = await state.get_data()
    pages = data['pages']
    new_pages = []
    for page in pages:
        new_pages.append([
            (item[0], item[1], (await rq.get_data_word_by_id(user_id, item[0], 'matching')).first())
            for item in page
        ])

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=new_pages)
    await state.update_data(less_matching=False)
    await display_words(callback, state)


# Удалить из отображения простые слова в категории
@router.callback_query(F.data.startswith('discard common'))
async def discard_common(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id
    category_id = callback.data.split('_')[1]

    # Получаем сложные слова для категории в зависимости от выбранных фильтров
    data = await state.get_data()
    less_name = data['less_name']
    less_matching = data['less_matching']
    
    if less_name == less_matching:
        items = [
            (item.id, item.name, item.matching)
            for item in await rq.get_difficult_words_by_category_id(user_id, category_id)
        ]
    elif less_name:
        items = [
            (item.id, item.matching)
            for item in await rq.get_difficult_words_by_category_id(user_id, category_id)
        ]
    else:
        items = [
            (item.id, item.name)
            for item in await rq.get_difficult_words_by_category_id(user_id, category_id)
        ]

    # Перемешиваем слова и разбиваем на страницы
    shuffle(items)
    pages = []
    len_items = len(items)
    for num in range((len_items + 14) // 15):
        pages.append(items[num * 15: (num + 1) * 15])

    # Обновляем состояние данных и вызываем обновление отображения
    await state.update_data(pages=pages)
    await state.update_data(current_page=0)
    await state.update_data(less_common_words=True)
    await display_words(callback, state)


# Восстановить отображение простых слов в категории
@router.callback_query(F.data.startswith('return common'))
async def return_common(callback: CallbackQuery, state: FSMContext):
    # Очищаем состояние и обновляем отображение
    await state.clear()
    await display_words(callback, state)


# Перемешать порядок слов в текущей странице слов
@router.callback_query(F.data.startswith('shuffle'))
async def shuffle_words(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id
    category_id = callback.data.split('_')[1]

    # Получаем данные состояния и текст текущего сообщения
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']
    less_name = data['less_name']
    less_matching = data['less_matching']
    less_common_words = data['less_common_words']

    # Перемешиваем строки слов

    lines = callback.message.text.split('\n')

    if len(lines) >= 4:
        name_category = ' '.join(lines[0].split()[1:])
        items = pages[current_page]
        last_item = items.pop(-1)
        shuffle(items)

        items.insert(0, last_item)
        new_line = '\n'

        def decorate(item, less_name, less_matching):
            if less_name == less_matching:
                name, matching = item
                return f"{name}  -  <i>{matching}</i>"
            if less_name:
                matching = item[0]
                return f"<i>{matching}</i>"
            name = item[0]
            return name

        new_content = f"Категория <b>{name_category}</b>\n\n{(new_line.join(decorate(item[1:], less_name, less_matching) for item in items))}"
        await callback.message.edit_text(
            new_content,
            reply_markup=await kb.inline_words(
                user_id=user_id,
                category_id=category_id,
                user_dict_id=(await rq.get_id_user_dict_by_id_category(user_id, category_id)).first(),
                current_page=current_page,
                cnt_pages=len(pages),
                is_not_empty=bool(pages[0] if pages else False),
                less_name=less_name,
                less_matching=less_matching,
                less_common_words=less_common_words
            ),
            parse_mode='html'
        )



#-----------------------------------------------------------------------------------
#endregion



#ДОБАВЛЕНИЕ НОВОГО СЛОВА В КАТЕГОРИЮ
#region
#-----------------------------------------------------------------------------------
#                /start/{название словаря}/{название категории}
#                /редактировать категорию/добавить новое слово
#-----------------------------------------------------------------------------------


# Начало процесса добавления нового слова
@router.callback_query(F.data.startswith('add new word'))
async def add_new_word(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id

    # Удаляем предыдущее сообщение и очищаем состояние
    await callback.message.delete()
    category_id = callback.data.split('_')[1]
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()

    if await state.get_state() is not None:
        await state.clear()

    # Переход в состояние добавления нового слова и обновление данных состояния
    await state.set_state(RegistationNewItem.name)
    await state.update_data(category_id=category_id, user_dict_id=user_dict_id)
    name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()

    await callback.message.answer(
        f'Введите новое слово категории <b>{name_category}</b>',
        reply_markup=await kb.reply_cancel_add_new_item(),
        parse_mode='html'
    )


# Получение слова от пользователя
@router.message(RegistationNewItem.name)
async def set_name_word(message: Message, state: FSMContext):

    user_id = message.from_user.id
    data = await state.get_data()
    category_id = data['category_id']

    # Если пользователь отменяет ввод
    if message.text == 'ОТМЕНА':
        await state.clear()
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        return await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )



    # Переход к получению перевода
    await state.update_data(name=message.text)
    await state.set_state(RegistationNewItem.matching)
    await message.answer('Введите перевод', reply_markup=await kb.reply_cancel_add_new_item())


# Получение перевода (соответствия) от пользователя
@router.message(RegistationNewItem.matching)
async def set_matching_word(message: Message, state: FSMContext):

    user_id = message.from_user.id

    await state.update_data(matching=message.text)
    data = await state.get_data()
    await state.clear()

    category_id = data['category_id']
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()

    # Проверка на существование слова в категории
    if (await rq.check_word_in_user_dict(user_id, data['name'], message.text, user_dict_id)).first() is not None:
        await state.clear()
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()
        await message.answer(f'Слово {data["name"]} - <i>{message.text}</i> уже есть в словаре {name_user_dict}', parse_mode='html')
        await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )
        return

    # Если пользователь отменяет ввод
    if message.text == 'ОТМЕНА':
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        return await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )

    # Добавление нового слова в категорию
    extra_words = await rq.add_new_word(user_id, category_id, data['name'], data['matching'])
    await message.answer(
        f"Слово <b>{data['name']}</b> - <i>{data['matching']}</i> добавлено!" + extra_words,
        parse_mode='html'
    )

    # Обновление и отображение страницы с категориями
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()
    await base_for_display_words(user_id, category_id, state)
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']
    less_name = data['less_name']
    less_matching = data['less_matching']
    less_common_words = data['less_common_words']

    # Отображение обновленного списка слов
    await message.answer(
        await text_words(
            user_id=user_id,
            category_id=category_id,
            pages=pages,
            page=current_page,
            less_name=less_name,
            less_matching=less_matching
        ),
        reply_markup=await kb.inline_words(
            user_id=user_id,
            category_id=category_id,
            user_dict_id=user_dict_id,
            current_page=current_page,
            cnt_pages=len(pages),
            is_not_empty=bool(pages[0] if pages else False),
            less_name=less_name,
            less_matching=less_matching,
            less_common_words=less_common_words
        ),
        parse_mode='html'
    )


#-----------------------------------------------------------------------------------
#endregion



#ДОБАВЛЕНИЕ НОВЫХ СЛОВ В КАТЕГОРИЮ
#region
#-----------------------------------------------------------------------------------
#                /start/{название словаря}/{название категории}
#                /редактировать категорию/добавить новые слова
#-----------------------------------------------------------------------------------


# Начало процесса добавления нескольких слов из файла
@router.callback_query(F.data.startswith('add words'))
async def add_new_words(callback: CallbackQuery, state: FSMContext):

    # Удаление сообщения и очищение состояния, если оно не пусто
    await callback.message.delete()
    if await state.get_state() is not None:
        await state.clear()

    # Установка состояния ожидания файла и обновление данных состояния
    await state.set_state(RegistationNewWords.file)
    await state.update_data(category_id=callback.data.split('_')[1])

    # Отправка инструкции для отправки файла с новыми словами
    await callback.message.answer(
        '''Отправьте файл текстового или табличного формата
Каждая строка текстового файла должна соответствовать шаблону: "<i>слово</i> - <i>перевод</i>"
Табличный файл должен быть без заголовков''',
        parse_mode='html',
        reply_markup=await kb.reply_cancel_add_new_item()
    )


# Обработка отправленного файла с новыми словами
@router.message(RegistationNewWords.file)
async def set_new_words(message: Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()
    category_id = data['category_id']
    await state.clear()

    # Если пользователь отменяет операцию
    if message.text == 'ОТМЕНА':
        user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )
        return

    # Проверка типа документа и его допустимости
    document = message.document
    if not document or document.mime_type == 'text/csv' or not (
        document.mime_type.startswith('text/') or
        document.mime_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']
    ):
        await message.answer("Пожалуйста, отправьте текстовый или XLSX документ")
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )
        return

    # Загрузка и обработка файла
    file = await message.bot.get_file(document.file_id)
    downloaded_file = await message.bot.download_file(file.file_path)

    try:
        # Чтение и разбор текстового файла
        if document.mime_type.startswith('text/'):
            content = downloaded_file.read().decode('utf-8')
            string_io = StringIO(content)
            words_data = [
                (name, matching)
                for name, matching in map(lambda line: line.rstrip().split(' - '), string_io)
            ]

        # Чтение и разбор табличного файла
        elif document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            excel_io = BytesIO(downloaded_file.read())
            workbook = load_workbook(excel_io, data_only=True)
            sheet = workbook.active
            words_data = [
                (name, matching)
                for name, matching in sheet.iter_rows(values_only=True)
            ]

        # Добавление слов в категорию
        extra_words = await rq.add_new_words(user_id, category_id, words_data)
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        await message.answer(f'Слова добавлены в категорию {name_category}!' + extra_words)

    except Exception as e:
        await message.answer("Произошла ошибка при обработке файла")

    # Обновление и отображение категории после добавления новых слов
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()
    await base_for_display_words(user_id, category_id, state)
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']
    less_name = data['less_name']
    less_matching = data['less_matching']
    less_common_words = data['less_common_words']


    await message.answer(
        await text_words(
            user_id=user_id,
            category_id=category_id,
            pages=pages,
            page=current_page,
            less_name=less_name,
            less_matching=less_matching
        ),
        reply_markup=await kb.inline_words(
            user_id=user_id,
            category_id=category_id,
            user_dict_id=user_dict_id,
            current_page=current_page,
            cnt_pages=len(pages),
            is_not_empty=bool(pages[0] if pages else False),
            less_name=less_name,
            less_matching=less_matching,
            less_common_words=less_common_words
        ),
        parse_mode='html'
    )


#endregion



#УДАЛЕНИЕ СЛОВА ИЗ КАТЕГОРИИ
#region
#-----------------------------------------------------------------------------------
#                 /start/{название словаря}/{название категории}
#                     /редактировать категорию/удалить слово
#-----------------------------------------------------------------------------------


# Начало процесса удаления слова
@router.callback_query(F.data.startswith('del word'))
async def del_word(callback: CallbackQuery, state: FSMContext):

    # Удаление сообщения и очистка состояния, если оно не пусто
    await callback.message.delete()
    if await state.get_state() is not None:
        await state.clear()

    # Установка состояния ожидания слова для удаления и обновление данных состояния
    await state.set_state(ConfirmDelWord.name)
    category_id = callback.data.split('_')[1]
    await state.update_data(category_id=category_id)

    # Отправка запроса на ввод слова для удаления
    await callback.message.answer(
        'Введите слово, которое хотите удалить',
        reply_markup=await kb.cancel_delete_word()
    )


# Получение слова для удаления
@router.message(ConfirmDelWord.name)
async def get_name_word_for_delete(message: Message, state: FSMContext):

    user_id = message.from_user.id
    name_word = message.text
    data = await state.get_data()
    category_id = data['category_id']
    await state.update_data(name=name_word)
    await state.set_state(ConfirmDelWord.matching)

    # Отмена удаления, возврат к категории
    if name_word == 'ОТМЕНА':
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )
        return

    # Отправка запроса на ввод слова для удаления
    await message.answer(
        'Введите соответствие слова',
        reply_markup=await kb.cancel_delete_word()
    )



# Получение соответствия для удаления
@router.message(ConfirmDelWord.matching)
async def get_matching_word_for_delete(message: Message, state: FSMContext):

    user_id = message.from_user.id
    matching_word = message.text
    data = await state.get_data()

    name_word = data['name']
    category_id = data['category_id']
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()

    # Отмена удаления, возврат к категории
    if matching_word == 'ОТМЕНА':
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )
        return

    # Проверка наличия слова в категории
    if (await rq.check_word_in_user_dict(user_id, name_word, matching_word, user_dict_id)).first() is None:
        name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
        await message.answer(f'Слово {name_word} - <i>{message.text}</i> отсутствует в категории {name_category}', parse_mode='html')
        await message.answer(
            f'Категория <b>{name_category}</b>',
            reply_markup=await kb.inline_edit_category(category_id),
            parse_mode='html'
        )
        return

    # Запрос на подтверждение удаления слова
    word_id = (await rq.get_id_word_by_name_and_matching(user_id, name_word, matching_word, category_id)).first()
    await message.answer(
        f'Вы точно хотите удалить слово {name_word} - <i>{message.text}</i>?',
        reply_markup=await kb.inline_confirm_del_word(word_id, category_id), 
        parse_mode='html'
    )



# Подтверждение удаления слова
@router.callback_query(F.data.startswith('confirm del word'))
async def confirm_del_word(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id
    word_id = callback.data.split('_')[1]
    name_word = (await rq.get_data_word_by_id(user_id, word_id, 'name')).first()
    category_id = callback.data.split('_')[2]
    name_category = (await rq.get_name_category_by_id(user_id, category_id)).first()
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()

    # Удаление слова из базы данных
    await rq.delete_word(user_id, word_id)
    await callback.message.edit_text(
        f"Слово <b>{name_word}</b> удалено из категории <b>{name_category}</b>",
        parse_mode='html'
    )

    # Очистка состояния, если оно активно
    if await state.get_state() is not None:
        await state.clear()

    # Обновление и отображение оставшихся слов категории
    await base_for_display_words(user_id, category_id, state)
    data = await state.get_data()
    pages = data['pages']
    current_page = data['current_page']
    less_name = data['less_name']
    less_matching = data['less_matching']
    less_common_words = data['less_common_words']


    await callback.message.answer(
        await text_words(
            user_id=user_id,
            category_id=category_id,
            pages=pages,
            page=current_page,
            less_name=less_name,
            less_matching=less_matching
        ),
        reply_markup=await kb.inline_words(
            user_id=user_id,
            category_id=category_id,
            user_dict_id=user_dict_id,
            current_page=current_page,
            cnt_pages=len(pages),
            is_not_empty=bool(pages[0] if pages else False),
            less_name=less_name,
            less_matching=less_matching,
            less_common_words=less_common_words
        ),
        parse_mode='html'
    )


#-----------------------------------------------------------------------------------
#endregion



#ИЗУЧЕНИЕ СЛОВ 
#region
#-----------------------------------------------------------------------------------
#                 /start/{название словаря}/учить слова или сложные слова
#-----------------------------------------------------------------------------------


# Обработка списка повторяемых слов 
@router.callback_query(F.data.startswith('learn repeating'))
async def list_of_repeating_words(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id
    await callback.message.delete()

    # Получение идентификаторов категорий
    user_dict_id = callback.data.split('_')[1]
    categories_id = {category_id for category_id in await(rq.get_categories_id_by_user_dict_id(user_id, user_dict_id))}

    # Формирование списка слов и перемешивание
    words = []
    for word in await rq.get_repeating_words_by_dict(user_id, user_dict_id):
        words.append({'id': word.id, 'name': word.name, 'matching': word.matching, 'level_difficulty': 0})
    shuffle(words)

    # Очистка и обновление состояния с добавлением данных
    if await state.get_state() is not None:
        await state.clear()
    await state.set_state(LearnWords.name)
    await state.update_data(categories_id=categories_id, user_dict_id=user_dict_id, words=words, order_difficult_words=[])

    await give_word(callback, state)


# Обработка списка обычных слов 
@router.callback_query(F.data.startswith('learn all'))
async def list_of_common_words(callback: CallbackQuery, state: FSMContext):

    user_id = callback.from_user.id

    # Получение идентификаторов категорий
    categories_id = {callback.data.split('_')[2],}
    category_id = categories_id.pop()
    categories_id.add(category_id)
    user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()

    # Формирование списка слов и перемешивание
    words = []
    for word in await rq.get_common_words_by_category_id(user_id, category_id):
        words.append({'id': word.id, 'name': word.name, 'matching': word.matching, 'level_difficulty': 0})
    shuffle(words)

    # Очистка и обновление состояния с добавлением данных
    if await state.get_state() is not None:
        await state.clear()
    await state.set_state(LearnWords.name)
    await state.update_data(categories_id=categories_id, user_dict_id=user_dict_id, words=words)

    await list_of_difficult_words(callback, state, True)


# Обработка списка сложных слов
@router.callback_query(F.data.startswith('learn diff'))
async def list_of_difficult_words(callback: CallbackQuery, state: FSMContext, is_call: bool = False):

    user_id = callback.from_user.id
    await callback.message.delete()

    # Обработка данных, если вызов через callback
    if not is_call:
        categories_id = {callback.data.split('_')[2],}
        category_id = categories_id.pop()
        categories_id.add(category_id)
        user_dict_id = (await rq.get_id_user_dict_by_id_category(user_id, category_id)).first()

        if await state.get_state() is not None:
            await state.clear()
        await state.set_state(LearnWords.name)
        await state.update_data(categories_id=categories_id, user_dict_id=user_dict_id, words=[])
    else:
        data = await state.get_data()
        categories_id = data['categories_id']
        category_id = categories_id.pop()
        categories_id.add(category_id)

    # Формирование списка сложных слов по уровням сложности
    difficult_words = {1: [], 2: [], 3: []}
    for word in await rq.get_difficult_words_by_category_id(user_id, category_id):
        difficult_words[word.level_difficulty].append((word.id, word.name, word.matching))

    # Перемешивание и создание очередности сложных слов
    tasks = []
    order_difficult_words = []
    for level in range(1, 4):
        shuffle(difficult_words[level])
        for word in difficult_words[level]:
            tasks.append(set_level_difficulty(user_id, {'id': word[0], 'name': word[1], 'matching': word[2], 'level_difficulty': level},
                                              order_difficult_words, categories_id))
    await asyncio.gather(*tasks)

    await state.update_data(order_difficult_words=order_difficult_words)
    await give_word(callback, state)


# Изменение уровня сложности слова
async def set_level_difficulty(user_id, word, order_difficult_words, categories_id: set, correct_answer=None):

    # Изменение уровня сложности в зависимости от правильного ответа
    if correct_answer is None:
        level_difficulty = word['level_difficulty']
    elif correct_answer:
        level_difficulty = word['level_difficulty'] + 1 if word['level_difficulty'] < 3 else 0
        await rq.set_new_level_difficulty_word(user_id, word['name'], level_difficulty, categories_id)
    else:
        level_difficulty = 1
        await rq.set_new_level_difficulty_word(user_id, word['name'], level_difficulty, categories_id)

    word['level_difficulty'] = level_difficulty

    # Добавление слова в очередь с интервалом по уровню сложности
    if level_difficulty != 0:
        step_for_word = 5 * level_difficulty
        for _ in range(step_for_word - 1 - len(order_difficult_words)):
            order_difficult_words.append(None)
        order_difficult_words.insert(5 * level_difficulty, word)


# Вывод слова из изучаемого раздела
async def give_word(message: Message, state: FSMContext):

    user_id = message.from_user.id

    data = await state.get_data()

    # Извлечение слова для показа, если список пуст, переход к следующему слову
    order_difficult_words = data['order_difficult_words']
    words = data['words']
    categories_id = data['categories_id']
    user_dict_id = data['user_dict_id']

    word = False
    if order_difficult_words:
        word = order_difficult_words.pop(0)
        if word is None:
            if words:
                word = words.pop(0)
            else:
                while word is None:
                    word = order_difficult_words.pop(0)
    elif words:
        word = words.pop(0)
        
    if isinstance(message, Message):
        send_message = message.answer
    else:
        send_message = message.message.answer
    
    if word:
        await send_message(word['matching'], reply_markup=await kb.reply_learn_word())
        await state.update_data(last_word=word)
    else:
        # Завершение обучения по всем словам

        if len(categories_id) <= 1:
            name_user_dict = (await rq.get_name_user_dict_by_id(user_id, user_dict_id)).first()
            extra_words = f'словаря <b>{name_user_dict}</b>'
        else:
            name_category = (await rq.get_name_category_by_id(user_id, categories_id.pop())).first()
            extra_words = f'категории <b>{name_category}</b>'
        
        await send_message(f'Поздравляю! Ты выучил слова {extra_words}', parse_mode='html')
        await cmd_start(message, state, new_user=False)


# Измененение дальнейшего повторения слова
async def set_new_data_for_repeat(user_id, word_id, is_correct_answer):
    word = (await rq.get_data_word_by_id(user_id, word_id, 'all') ).first()

    is_repeating = word.is_repeating
    date_for_repeat = word.date_for_repeat
    repeating_interval = word.repeating_interval

    if is_correct_answer:
        if is_repeating:
            today = date.today()
            await rq.word_was_answered_today(user_id, word_id, today)

            if date_for_repeat <= today:
                repeating_interval = min(repeating_interval * 2, 32)

            new_date_answer = today + timedelta(days=repeating_interval)
            await rq.set_new_date_answer(user_id, word_id, new_date_answer, repeating_interval)
    else:
        if is_repeating:
            today = date.today()
            await rq.word_was_answered_today(user_id, word_id, today)

            if date_for_repeat <= today:
                repeating_interval = max(repeating_interval // 2, 1)
                
            new_date_answer = today + timedelta(days=repeating_interval) 
            await rq.set_new_date_answer(user_id, word_id, new_date_answer, repeating_interval)

        else:
            today = date.today()
            new_date_answer = today + timedelta(days=1) 
            await rq.make_word_repeating(user_id, word_id, today, new_date_answer)


# Проверка правильности перевода
@router.message(LearnWords.name)
async def get_name(message: Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()

    # Проверка ответа пользователя на соответствие
    word = data['last_word']
    order_difficult_words = data['order_difficult_words']
    if message.text != 'ЗАКОНЧИТЬ':
        is_correct_answer = (message.text.lower() == word['name'].lower())
        if not is_correct_answer:
            await message.answer(f"{word['name']}  -  {word['matching']}")

        # Обновление уровня сложности слова, если ответ неверный или слово сложное
        if (not is_correct_answer) or (word['level_difficulty'] != 0):
            await set_level_difficulty(user_id, word, order_difficult_words, data['categories_id'], is_correct_answer)

        await set_new_data_for_repeat(user_id, word['id'], is_correct_answer)

        await give_word(message, state)

    else:
        await state.clear()
        await message.answer('Ты большой молодец!')

        await cmd_start(message, state, new_user=False)


#-----------------------------------------------------------------------------------
#endregion



#endregion

